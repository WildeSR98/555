"""
Создание структуры папок на сетевом диске при создании проекта.
Вызывается из FastAPI (projects_api.py) после успешного db.commit().

Структура (зависит от типа этапа):
  FULL  → NET_ROOT / project_name / stage / part_number / serial_number
  PN    → NET_ROOT / project_name / stage / part_number
  EMPTY → NET_ROOT / project_name / stage   (пустая папка)

Env:
  NET_PROJECTS_DIR  — корневая папка на сетевом диске
  NET_STAGES_FULL   — этапы с полной вложенностью PN/SN (через запятую)
  NET_STAGES_PN     — этапы только с PN (без SN)  (через запятую)
  NET_STAGES_EMPTY  — пустые папки этапов          (через запятую)
"""

import os
import logging
from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(BASE_DIR / '.env', override=True)

NET_ROOT = Path(os.getenv('NET_PROJECTS_DIR', r'\\192.168.106.29\PR_DEP'))

STAGES_FULL = [
    s.strip() for s in
    os.getenv('NET_STAGES_FULL', 'Complectation,OTK,Packing Stand,Tests').split(',')
    if s.strip()
]
STAGES_PN = [
    s.strip() for s in
    os.getenv('NET_STAGES_PN', 'Vibrostand').split(',')
    if s.strip()
]
STAGES_EMPTY = [
    s.strip() for s in
    os.getenv('NET_STAGES_EMPTY', 'FRU,Accounting,Warehouse').split(',')
    if s.strip()
]

log = logging.getLogger(__name__)


def create_project_folders(project_name: str, devices: list) -> dict:
    """
    Создаёт структуру папок на сетевом диске для нового проекта.

    Args:
        project_name: имя проекта (используется как имя корневой папки)
        devices:      список словарей [{part_number: str, serial_number: str}]

    Returns:
        {'ok': bool, 'path': str, 'created': int, 'error': str | None}
    """
    project_root = NET_ROOT / project_name
    result = {'ok': False, 'path': str(project_root), 'created': 0, 'error': None}

    # Группируем серийники по партномеру
    pn_map: dict[str, list[str]] = {}
    for dev in devices:
        pn = dev.get('part_number') or 'NO_PN'
        sn = dev.get('serial_number') or 'NO_SN'
        pn_map.setdefault(pn, []).append(sn)

    # Создаём корневую папку проекта
    try:
        project_root.mkdir(parents=True, exist_ok=True)
        log.info(f'[NET] Корень: {project_root}')
    except Exception as e:
        result['error'] = f'Не удалось создать {project_root}: {e}'
        log.error(result['error'])
        return result

    def _mkdir(p: Path) -> None:
        p.mkdir(exist_ok=True)
        result['created'] += 1

    # ── FULL: этап / PN / SN ─────────────────────────────────────────────────
    for stage in STAGES_FULL:
        stage_path = project_root / stage
        _mkdir(stage_path)
        log.info(f'[NET]  ├─ {stage}/')
        for pn, sns in pn_map.items():
            pn_path = stage_path / pn
            _mkdir(pn_path)
            log.info(f'[NET]  │   ├─ {pn}/')
            for sn in sns:
                try:
                    _mkdir(pn_path / sn)
                    log.info(f'[NET]  │   │   └─ {sn}/')
                except Exception as e:
                    log.warning(f'[NET]  │   │   └─ ОШИБКА {sn}: {e}')

    # ── PN only: этап / PN (без SN) ──────────────────────────────────────────
    for stage in STAGES_PN:
        stage_path = project_root / stage
        _mkdir(stage_path)
        log.info(f'[NET]  ├─ {stage}/')
        for pn in pn_map:
            try:
                _mkdir(stage_path / pn)
                log.info(f'[NET]  │   └─ {pn}/')
            except Exception as e:
                log.warning(f'[NET]  │   └─ ОШИБКА {pn}: {e}')

    # ── EMPTY: только сама папка этапа ───────────────────────────────────────
    for stage in STAGES_EMPTY:
        try:
            _mkdir(project_root / stage)
            log.info(f'[NET]  ├─ {stage}/ (пусто)')
        except Exception as e:
            log.warning(f'[NET]  ├─ ОШИБКА {stage}: {e}')

    result['ok'] = True
    log.info(f'[NET] Готово: создано папок={result["created"]} в {project_root}')
    return result


def create_project_excel(project_name: str, devices: list) -> dict:
    """
    Создаёт Excel-файл {project_name}_mac_sn.xlsx в корне проекта на сетевом диске.

    Args:
        project_name: имя проекта
        devices: список словарей {part_number, serial_number, mac1, mac2, category}

    Returns:
        {'ok': bool, 'path': str, 'error': str | None}
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return {'ok': False, 'path': '', 'error': 'openpyxl не установлен. pip install openpyxl'}

    excel_path = NET_ROOT / project_name / f"{project_name}_mac_sn.xlsx"
    result = {'ok': False, 'path': str(excel_path), 'error': None}

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Devices'

        header_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill  = PatternFill('solid', fgColor='2F5496')
        center_align = Alignment(horizontal='center', vertical='center')
        data_font    = Font(name='Consolas', size=10)
        thin         = Side(style='thin')
        thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
        alt_fill     = PatternFill('solid', fgColor='DCE6F1')

        headers    = ['PN', 'SN', 'MAC1 (LAN)', 'MAC2 (iDRAC/BMC)']
        col_widths = [22, 22, 22, 22]

        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center_align
            cell.border    = thin_border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[1].height = 20

        sorted_devices = sorted(devices, key=lambda d: (d.get('part_number', ''), d.get('serial_number', '')))

        for row_idx, dev in enumerate(sorted_devices, start=2):
            fill = alt_fill if row_idx % 2 == 0 else None
            values = [
                dev.get('part_number', ''),
                dev.get('serial_number', ''),
                dev.get('mac1', ''),
                dev.get('mac2', ''),
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.font      = data_font
                cell.alignment = center_align
                cell.border    = thin_border
                if fill:
                    cell.fill = fill

        ws.freeze_panes = 'A2'
        wb.save(excel_path)
        log.info(f'[NET] Excel сохранён: {excel_path}')
        result['ok'] = True

    except Exception as e:
        result['error'] = str(e)
        log.error(f'[NET] Ошибка создания Excel: {e}')

    return result

# ── Запуск напрямую для проверки ─────────────────────────────────────────────
if __name__ == '__main__':
    import argparse, json, sys
    logging.basicConfig(level=logging.INFO, format='%(message)s')

    parser = argparse.ArgumentParser(description='Создать структуру папок проекта на сетевом диске')
    parser.add_argument('--name',    required=True, help='Имя проекта (папка)')
    parser.add_argument('--devices', default='[]',  help='JSON: [{"part_number":"PN-001","serial_number":"SN00001"}]')
    args = parser.parse_args()

    devs = json.loads(args.devices)
    res  = create_project_folders(args.name, devs)
    print(json.dumps(res, ensure_ascii=False, indent=2))
    sys.exit(0 if res['ok'] else 1)
